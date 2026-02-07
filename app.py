"""
FNCE102 Portfolio Analyzer - Web Interface
A Flask-based web dashboard for interactive portfolio analysis
"""

from flask import Flask, render_template, request, jsonify, send_from_directory
import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt

# Import our modules
from data_fetcher import fetch_data, calculate_returns
from portfolio import Portfolio, create_benchmark_portfolio
from metrics import PerformanceMetrics, compare_portfolios
from report import create_summary_report, export_to_excel

app = Flask(__name__, static_folder='static', template_folder='templates')

# Create necessary directories for deployment (Render/Gunicorn)
os.makedirs('output', exist_ok=True)
os.makedirs('static', exist_ok=True)
os.makedirs('templates', exist_ok=True)

# Global cache for data
DATA_CACHE = {}

# Common ETFs organized by category
COMMON_ETFS = {
    "US Equities": [
        {"ticker": "VTI", "name": "Vanguard Total Stock Market"},
        {"ticker": "SPY", "name": "SPDR S&P 500"},
        {"ticker": "QQQ", "name": "Invesco Nasdaq-100"},
        {"ticker": "VGT", "name": "Vanguard Information Technology"},
        {"ticker": "VNQ", "name": "Vanguard Real Estate"},
        {"ticker": "IWM", "name": "iShares Russell 2000 Small Cap"},
        {"ticker": "XLF", "name": "Financial Select Sector"},
        {"ticker": "XLE", "name": "Energy Select Sector"},
        {"ticker": "XLV", "name": "Health Care Select Sector"},
        {"ticker": "XLK", "name": "Technology Select Sector"},
    ],
    "International Equities": [
        {"ticker": "VEA", "name": "Vanguard FTSE Developed Markets"},
        {"ticker": "VWO", "name": "Vanguard FTSE Emerging Markets"},
        {"ticker": "EFA", "name": "iShares MSCI EAFE"},
        {"ticker": "EEM", "name": "iShares MSCI Emerging Markets"},
    ],
    "Bonds": [
        {"ticker": "BND", "name": "Vanguard Total Bond Market"},
        {"ticker": "AGG", "name": "iShares Core US Aggregate Bond"},
        {"ticker": "LQD", "name": "iShares Investment Grade Corporate"},
        {"ticker": "TIP", "name": "iShares TIPS Bond"},
        {"ticker": "SHY", "name": "iShares 1-3 Year Treasury"},
        {"ticker": "TLT", "name": "iShares 20+ Year Treasury"},
        {"ticker": "HYG", "name": "iShares High Yield Corporate"},
    ],
    "Alternatives": [
        {"ticker": "GLD", "name": "SPDR Gold Shares"},
        {"ticker": "SLV", "name": "iShares Silver Trust"},
        {"ticker": "DBC", "name": "Invesco DB Commodity Index"},
        {"ticker": "VNQ", "name": "Vanguard Real Estate"},
    ]
}

CLIENT_PROFILES = {
    "conservative": {
        "name": "Richard Tan",
        "description": "Conservative - Capital Preservation & Income",
        "objectives": ["Max Drawdown < 10%", "Lower volatility than SPY"],
        "suggested": {"BND": 0.30, "LQD": 0.15, "TIP": 0.10, "VTI": 0.20, "VEA": 0.10, "GLD": 0.10, "SHY": 0.05}
    },
    "balanced": {
        "name": "Sophia Lim", 
        "description": "Balanced - Growth with Moderate Risk",
        "objectives": ["Max Drawdown < 20%", "Volatility < 15%"],
        "suggested": {"VTI": 0.25, "VGT": 0.10, "VEA": 0.15, "VWO": 0.05, "VNQ": 0.05, "BND": 0.20, "LQD": 0.10, "GLD": 0.05, "DBC": 0.05}
    },
    "aggressive": {
        "name": "David Lee",
        "description": "Aggressive - High Growth & Alternatives", 
        "objectives": ["Sharpe Ratio > 1", "Target 2x SPY returns"],
        "suggested": {"VTI": 0.20, "QQQ": 0.20, "VGT": 0.15, "VWO": 0.10, "IWM": 0.10, "VNQ": 0.05, "GLD": 0.05, "BND": 0.10, "DBC": 0.05}
    }
}


@app.route('/')
def index():
    return render_template('index.html', 
                         etfs=COMMON_ETFS, 
                         profiles=CLIENT_PROFILES)


@app.route('/api/analyze', methods=['POST'])
def analyze():
    """Run portfolio analysis with given allocations."""
    try:
        data = request.json
        allocations = data.get('allocations', {})
        profile = data.get('profile', 'balanced')
        start_date = data.get('start_date', '2009-12-01')
        end_date = data.get('end_date', '2025-12-31')
        
        # Validate allocations
        if not allocations:
            return jsonify({"error": "No allocations provided"}), 400
        
        total = sum(allocations.values())
        if not np.isclose(total, 1.0, atol=0.02):
            return jsonify({"error": f"Allocations must sum to 100%. Current: {total*100:.1f}%"}), 400
        
        # Get tickers including SPY for benchmark
        tickers = list(allocations.keys()) + ["SPY"]
        tickers = list(set(tickers))
        
        # Fetch data (with caching)
        cache_key = f"{','.join(sorted(tickers))}_{start_date}_{end_date}"
        if cache_key in DATA_CACHE:
            prices = DATA_CACHE[cache_key]
        else:
            prices = fetch_data(tickers, start_date, end_date)
            DATA_CACHE[cache_key] = prices
        
        # Calculate returns
        monthly_returns = calculate_returns(prices, frequency="monthly")
        
        # Create portfolio
        portfolio = Portfolio(
            name="Custom Portfolio",
            allocations=allocations,
            returns_data=monthly_returns
        )
        
        benchmark = create_benchmark_portfolio(monthly_returns, "SPY")
        
        # Calculate metrics
        metrics_calc = PerformanceMetrics(
            portfolio.portfolio_returns,
            benchmark.portfolio_returns,
            risk_free_rate=0.02
        )
        
        portfolio_metrics = metrics_calc.get_all_metrics()
        
        # Get benchmark metrics
        bench_metrics = PerformanceMetrics(
            benchmark.portfolio_returns,
            benchmark.portfolio_returns,
            risk_free_rate=0.02
        ).get_all_metrics()
        
        # Calculate cumulative returns for chart
        port_cum = (1 + portfolio.portfolio_returns).cumprod()
        bench_cum = (1 + benchmark.portfolio_returns).cumprod()
        
        # Prepare chart data  
        dates = [d.strftime('%Y-%m') for d in port_cum.index]
        
        # Annual returns
        port_annual = (1 + portfolio.portfolio_returns).resample('YE').prod() - 1
        bench_annual = (1 + benchmark.portfolio_returns).resample('YE').prod() - 1
        
        # Check objectives
        objectives_met = check_objectives(profile, portfolio_metrics)
        
        # Format metrics for display
        def format_metric(key, value):
            if isinstance(value, pd.Timestamp):
                return value.strftime('%Y-%m-%d')
            elif key in ["Maximum Drawdown", "Annualized Return", "Annualized Volatility", "Jensen's Alpha"]:
                return f"{value*100:.2f}%"
            elif isinstance(value, float):
                return f"{value:.4f}"
            return str(value)
        
        response = {
            "success": True,
            "portfolio_metrics": {k: format_metric(k, v) for k, v in portfolio_metrics.items() 
                                  if not isinstance(v, pd.Timestamp) or k.startswith("Max DD")},
            "benchmark_metrics": {k: format_metric(k, v) for k, v in bench_metrics.items()
                                  if not isinstance(v, pd.Timestamp) or k.startswith("Max DD")},
            "cumulative_returns": {
                "dates": dates,
                "portfolio": [round(v, 4) for v in port_cum.values.tolist()],
                "benchmark": [round(v, 4) for v in bench_cum.values.tolist()]
            },
            "annual_returns": {
                "years": [int(d.year) for d in port_annual.index],
                "portfolio": [round(v*100, 2) for v in port_annual.values.tolist()],
                "benchmark": [round(v*100, 2) for v in bench_annual.values.tolist()]
            },
            "objectives": objectives_met,
            "data_period": f"{monthly_returns.index[0].strftime('%Y-%m')} to {monthly_returns.index[-1].strftime('%Y-%m')}",
            "months": len(monthly_returns)
        }
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def check_objectives(profile, metrics):
    """Check if portfolio meets client objectives."""
    objectives = []
    
    ann_return = float(metrics.get("Annualized Return", 0))
    ann_vol = float(metrics.get("Annualized Volatility", 0))
    max_dd = float(abs(metrics.get("Maximum Drawdown", 0)))
    sharpe = float(metrics.get("Sharpe Ratio", 0))
    
    if profile == "conservative":
        objectives.append({
            "name": "Max Drawdown < 10%",
            "target": 0.10,
            "actual": max_dd,
            "met": bool(max_dd < 0.10)
        })
        objectives.append({
            "name": "Capital Preservation",
            "target": "Steady",
            "actual": f"{ann_vol*100:.1f}% vol",
            "met": bool(ann_vol < 0.10)
        })
    elif profile == "balanced":
        objectives.append({
            "name": "Max Drawdown < 20%",
            "target": 0.20,
            "actual": max_dd,
            "met": bool(max_dd < 0.20)
        })
        objectives.append({
            "name": "Volatility < 15%",
            "target": 0.15,
            "actual": ann_vol,
            "met": bool(ann_vol < 0.15)
        })
    elif profile == "aggressive":
        objectives.append({
            "name": "Sharpe Ratio > 1",
            "target": 1.0,
            "actual": sharpe,
            "met": bool(sharpe > 1.0)
        })
        objectives.append({
            "name": "High Growth",
            "target": "> SPY",
            "actual": f"{ann_return*100:.1f}%",
            "met": bool(ann_return > 0.12)
        })
    
    return objectives


@app.route('/api/export', methods=['POST'])
def export():
    """Export analysis to Excel."""
    try:
        data = request.json
        allocations = data.get('allocations', {})
        profile = data.get('profile', 'balanced')
        
        if not allocations:
            return jsonify({"error": "No allocations provided. Please run analysis first."}), 400
        
        # Run analysis - dates match user's example code
        tickers = list(allocations.keys()) + ["SPY"]
        prices = fetch_data(list(set(tickers)), "2009-12-01", "2025-12-31")
        monthly_returns = calculate_returns(prices, frequency="monthly")
        
        portfolio = Portfolio(name="Custom Portfolio", allocations=allocations, returns_data=monthly_returns)
        benchmark = create_benchmark_portfolio(monthly_returns, "SPY")
        
        metrics_calc = PerformanceMetrics(portfolio.portfolio_returns, benchmark.portfolio_returns)
        port_metrics = metrics_calc.get_all_metrics()
        
        bench_metrics = PerformanceMetrics(
            benchmark.portfolio_returns,
            benchmark.portfolio_returns,
            risk_free_rate=0.02
        ).get_all_metrics()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"portfolio_report_{timestamp}.xlsx"
        output_path = os.path.join("output", filename)
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Raw Daily Close Prices with Daily Returns
            # Format: date | ticker1 | ticker1_return | ticker2 | ticker2_return | ...
            user_tickers = list(allocations.keys())
            all_tickers = user_tickers + ["SPY"]
            
            # Get daily close prices
            daily_close_df = prices[all_tickers].copy()
            
            # Calculate daily returns for each ticker
            daily_returns = daily_close_df.pct_change()
            
            # Create combined dataframe with Price and Return columns side by side
            combined_data = pd.DataFrame(index=daily_close_df.index)
            combined_data.index.name = 'date'
            
            for ticker in all_tickers:
                # Add close price column
                combined_data[ticker] = daily_close_df[ticker].round(2)
                # Add daily return column (as percentage)
                combined_data[f'{ticker}_Return%'] = (daily_returns[ticker] * 100).round(4)
            
            combined_data.to_excel(writer, sheet_name="Daily Data", index=True)
            
            # Sheet 2: Monthly Returns (format matches user's example code)
            # monthly_returns = data.resample('ME').ffill().pct_change().dropna()
            monthly_returns_df = monthly_returns[all_tickers].copy()
            monthly_returns_df.index.name = 'date'
            # Keep as decimal format (like user's csv output)
            monthly_returns_df = monthly_returns_df.round(6)
            monthly_returns_df.to_excel(writer, sheet_name="Monthly Returns", index=True)
            
            # Combined Summary Sheet (will be formatted with openpyxl after)
            # Create placeholder - we'll add content via openpyxl for better formatting
            pd.DataFrame({"Summary": ["See formatted sections below"]}).to_excel(writer, sheet_name="Summary", index=False)
            
            # Store data for Summary sheet formatting later
            summary_info = {
                'profile': profile,
                'allocations': allocations,
                'port_metrics': port_metrics,
                'bench_metrics': bench_metrics,
                'port_annual': (1 + portfolio.portfolio_returns).resample('YE').prod() - 1,
                'bench_annual': (1 + benchmark.portfolio_returns).resample('YE').prod() - 1,
                'n_months': len(monthly_returns)
            }
            
            # Sheet 6: Cumulative Growth
            port_cum = (1 + portfolio.portfolio_returns).cumprod()
            bench_cum = (1 + benchmark.portfolio_returns).cumprod()
            cum_df = pd.DataFrame({
                "Date": port_cum.index.strftime("%Y-%m"),
                "Portfolio ($1 Growth)": port_cum.round(4).values,
                "SPY Benchmark ($1 Growth)": bench_cum.round(4).values
            })
            cum_df.to_excel(writer, sheet_name="Cumulative Growth", index=False)
            
            # ========== EFFICIENT FRONTIER WORKINGS WITH EXCEL FORMULAS ==========
            # Use monthly returns for efficient frontier calculations
            ef_returns = monthly_returns[user_tickers].copy()
            n_months = len(ef_returns)
            
            # Pre-calculate values for static sheets (we'll add formula sheets separately)
            mean_monthly = ef_returns.mean()
            mean_annual = mean_monthly * 12
            std_monthly = ef_returns.std()
            std_annual = std_monthly * np.sqrt(12)
            
            stats_df = pd.DataFrame({
                'Ticker': user_tickers,
                'Mean Monthly Return': mean_monthly.values,
                'Mean Annual Return': mean_annual.values,
                'Monthly Std Dev': std_monthly.values,
                'Annual Std Dev (Volatility)': std_annual.values,
                'Sharpe Ratio (Rf=2%)': ((mean_annual - 0.02) / std_annual).values
            })
            stats_df.to_excel(writer, sheet_name="Asset Statistics", index=False)
            
            # Correlation Matrix
            corr_matrix = ef_returns.corr()
            corr_matrix.to_excel(writer, sheet_name="Correlation Matrix", index=True)
            
            # Covariance Matrix (Annualized)
            cov_matrix = ef_returns.cov() * 12
            cov_matrix.to_excel(writer, sheet_name="Covariance Matrix", index=True)
            
            # Sheet 10: Efficient Frontier Calculation
            # Generate efficient frontier portfolios
            n_assets = len(user_tickers)
            n_portfolios = 50  # Number of points on frontier
            
            # Arrays to store results
            ef_results = []
            
            # Risk-free rate
            rf = 0.02
            
            # Generate random portfolios for Monte Carlo (5,000 trials)
            np.random.seed(42)
            n_random = 5000
            
            for _ in range(n_random):
                weights = np.random.random(n_assets)
                weights = weights / np.sum(weights)
                
                port_return = np.sum(weights * mean_annual)
                port_std = np.sqrt(np.dot(weights.T, np.dot(cov_matrix.values, weights)))
                sharpe = (port_return - rf) / port_std if port_std > 0 else 0
                
                # Build row with weights first (like in user's example)
                row_data = {ticker: w for ticker, w in zip(user_tickers, weights)}
                row_data['Portfolio Std Dev'] = port_std
                row_data['Portfolio Expected Returns'] = port_return
                row_data['Sharpe Ratio'] = sharpe
                
                ef_results.append(row_data)
            
            ef_df = pd.DataFrame(ef_results)
            
            # Reorder columns: tickers first, then metrics
            weight_cols = user_tickers
            metric_cols = ['Portfolio Std Dev', 'Portfolio Expected Returns', 'Sharpe Ratio']
            ef_df = ef_df[weight_cols + metric_cols]
            
            # Find optimal portfolios
            max_sharpe_idx = ef_df['Sharpe Ratio'].idxmax()
            min_vol_idx = ef_df['Portfolio Std Dev'].idxmin()
            
            max_sharpe_portfolio = ef_df.loc[max_sharpe_idx]
            min_vol_portfolio = ef_df.loc[min_vol_idx]
            
            # Export ALL points with weights for proper EF visualization in Excel
            ef_df.to_excel(writer, sheet_name="Efficient Frontier", index=False)
            
            # Sheet 11: Optimal Portfolios
            optimal_data = []
            
            # Max Sharpe Portfolio
            optimal_data.append({
                'Portfolio Type': 'Maximum Sharpe Ratio',
                'Expected Return': f"{max_sharpe_portfolio['Portfolio Expected Returns']*100:.2f}%",
                'Volatility': f"{max_sharpe_portfolio['Portfolio Std Dev']*100:.2f}%",
                'Sharpe Ratio': f"{max_sharpe_portfolio['Sharpe Ratio']:.4f}"
            })
            
            # Min Volatility Portfolio
            optimal_data.append({
                'Portfolio Type': 'Minimum Volatility',
                'Expected Return': f"{min_vol_portfolio['Portfolio Expected Returns']*100:.2f}%",
                'Volatility': f"{min_vol_portfolio['Portfolio Std Dev']*100:.2f}%",
                'Sharpe Ratio': f"{min_vol_portfolio['Sharpe Ratio']:.4f}"
            })
            
            # Current Portfolio
            current_weights = np.array([allocations.get(t, 0) for t in user_tickers])
            current_return = np.sum(current_weights * mean_annual)
            current_std = np.sqrt(np.dot(current_weights.T, np.dot(cov_matrix.values, current_weights)))
            current_sharpe = (current_return - rf) / current_std if current_std > 0 else 0
            
            optimal_data.append({
                'Portfolio Type': 'Your Current Portfolio',
                'Expected Return': f"{current_return*100:.2f}%",
                'Volatility': f"{current_std*100:.2f}%",
                'Sharpe Ratio': f"{current_sharpe:.4f}"
            })
            
            pd.DataFrame(optimal_data).to_excel(writer, sheet_name="Optimal Portfolios", index=False)
            
            # Sheet 12: Optimal Weights
            weights_data = {'Ticker': user_tickers}
            weights_data['Max Sharpe Weights'] = [f"{max_sharpe_portfolio[t]*100:.2f}%" for t in user_tickers]
            weights_data['Min Vol Weights'] = [f"{min_vol_portfolio[t]*100:.2f}%" for t in user_tickers]
            weights_data['Your Current Weights'] = [f"{allocations.get(t, 0)*100:.2f}%" for t in user_tickers]
            
            pd.DataFrame(weights_data).to_excel(writer, sheet_name="Optimal Weights", index=False)
            
            # Sheet 13: Efficient Frontier Workings Explanation
            workings_text = [
                ["EFFICIENT FRONTIER CALCULATION WORKINGS"],
                [""],
                ["Step 1: Calculate Asset Statistics"],
                ["- Mean Monthly Return = Average of monthly returns"],
                ["- Mean Annual Return = Monthly Mean × 12"],
                ["- Monthly Std Dev = Standard deviation of monthly returns"],
                ["- Annual Volatility = Monthly Std Dev × √12"],
                [""],
                ["Step 2: Calculate Covariance Matrix"],
                ["- Monthly Covariance = Cov(R_i, R_j) for all asset pairs"],
                ["- Annual Covariance = Monthly Cov × 12"],
                [""],
                ["Step 3: Generate Random Portfolios (Monte Carlo)"],
                ["- Generated 5,000 random weight combinations"],
                ["- For each portfolio:"],
                ["  - Portfolio Return = Σ(weight_i × return_i)"],
                ["  - Portfolio Variance = w' × Σ × w (where Σ is covariance matrix)"],
                ["  - Portfolio Volatility = √Variance"],
                ["  - Sharpe Ratio = (Return - Rf) / Volatility, where Rf = 2%"],
                [""],
                ["Step 4: Find Optimal Portfolios"],
                ["- Maximum Sharpe: Portfolio with highest Sharpe Ratio"],
                ["- Minimum Volatility: Portfolio with lowest standard deviation"],
                [""],
                ["Formula References:"],
                ["- Portfolio Return: E(Rp) = Σ w_i × E(R_i)"],
                ["- Portfolio Variance: σ²p = ΣΣ w_i × w_j × σ_ij"],
                ["- Sharpe Ratio: SR = (E(Rp) - Rf) / σp"],
            ]
            workings_df = pd.DataFrame(workings_text, columns=["Efficient Frontier Methodology"])
            workings_df.to_excel(writer, sheet_name="EF Workings", index=False)
        
        # Generate Efficient Frontier Chart
        chart_filename = f"efficient_frontier_{timestamp}.png"
        chart_path = os.path.join("output", chart_filename)
        
        # Create the chart
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Plot all random portfolios (scatter)
        # Plot all random portfolios (scatter)
        scatter = ax.scatter(
            ef_df['Portfolio Std Dev'] * 100, 
            ef_df['Portfolio Expected Returns'] * 100,
            c='#4B87B9',  # Blue color
            alpha=0.6,
            s=15,
            label='Random Portfolios'
        )
        
        # Add colorbar for Sharpe ratio
        # Removed colorbar as points are now single color
        # cbar = plt.colorbar(scatter, ax=ax)
        # cbar.set_label('Sharpe Ratio', fontsize=12)
        
        # ===== EFFICIENT FRONTIER CURVE (REMOVED) =====
        # User requested to remove the curve line

        
        # ===== CAPITAL ALLOCATION LINE (CAL) =====
        # CAL goes from risk-free rate (0, Rf) through tangency portfolio (Max Sharpe)
        rf_rate = 0.02  # 2% risk-free rate
        tangency_vol = max_sharpe_portfolio['Portfolio Std Dev']
        tangency_ret = max_sharpe_portfolio['Portfolio Expected Returns']
        
        # CAL slope = Sharpe ratio of tangency portfolio
        cal_slope = (tangency_ret - rf_rate) / tangency_vol
        
        # Extend CAL from 0 to beyond max volatility
        cal_vol_range = np.linspace(0, ef_df['Portfolio Std Dev'].max() * 1.2, 100)
        cal_returns = rf_rate + cal_slope * cal_vol_range
        
        ax.plot(
            cal_vol_range * 100,
            cal_returns * 100,
            color='orange',
            linewidth=2.5,
            linestyle='--',
            label=f'CAL (slope={cal_slope:.2f})',
            zorder=3
        )
        
        # Mark risk-free rate point
        ax.scatter(0, rf_rate * 100, marker='o', color='gold', s=150, 
                   edgecolors='black', linewidths=1, label=f'Risk-Free Rate ({rf_rate*100:.0f}%)', zorder=5)
        
        # Plot Max Sharpe portfolio (star)
        ax.scatter(
            max_sharpe_portfolio['Portfolio Std Dev'] * 100,
            max_sharpe_portfolio['Portfolio Expected Returns'] * 100,
            marker='*',
            color='red',
            s=500,
            edgecolors='black',
            linewidths=1,
            label=f"Tangency Portfolio (SR={max_sharpe_portfolio['Sharpe Ratio']:.2f})",
            zorder=5
        )
        
        # Plot Min Volatility portfolio (diamond)
        ax.scatter(
            min_vol_portfolio['Portfolio Std Dev'] * 100,
            min_vol_portfolio['Portfolio Expected Returns'] * 100,
            marker='D',
            color='blue',
            s=200,
            edgecolors='black',
            linewidths=1,
            label=f"Min Volatility (σ={min_vol_portfolio['Portfolio Std Dev']*100:.1f}%)",
            zorder=5
        )
        
        # Plot Current Portfolio (square)
        ax.scatter(
            current_std * 100,
            current_return * 100,
            marker='s',
            color='green',
            s=200,
            edgecolors='black',
            linewidths=1,
            label=f"Your Portfolio (SR={current_sharpe:.2f})",
            zorder=5
        )
        
        # Labels and title
        ax.set_xlabel('Volatility (Annual %)', fontsize=14)
        ax.set_ylabel('Expected Return (Annual %)', fontsize=14)
        ax.set_title('Efficient Frontier with Capital Allocation Line\nPortfolio Optimization', fontsize=16, fontweight='bold')
        ax.legend(loc='upper left', fontsize=9, framealpha=0.9)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(left=0)
        ax.set_ylim(bottom=0)
        
        # Add text annotation for key portfolios
        ax.annotate(
            'Tangency\n(Max Sharpe)',
            xy=(max_sharpe_portfolio['Portfolio Std Dev'] * 100, max_sharpe_portfolio['Portfolio Expected Returns'] * 100),
            xytext=(15, 15),
            textcoords='offset points',
            fontsize=9,
            fontweight='bold',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
        )
        
        ax.annotate(
            'Min Vol',
            xy=(min_vol_portfolio['Portfolio Std Dev'] * 100, min_vol_portfolio['Portfolio Expected Returns'] * 100),
            xytext=(10, -20),
            textcoords='offset points',
            fontsize=9,
            fontweight='bold',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7)
        )
        
        plt.tight_layout()
        plt.savefig(chart_path, dpi=150, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        # Embed chart and add formula-based sheets using openpyxl
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = load_workbook(output_path)
        
        # ========== FORMAT CONSOLIDATED SUMMARY SHEET ==========
        ws_summary = wb["Summary"]
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=12, color="1F4E79")
        
        # Clear placeholder content
        ws_summary.delete_rows(1, 2)
        
        row = 1
        
        # === SECTION 1: Portfolio Overview ===
        ws_summary.cell(row=row, column=1, value="PORTFOLIO OVERVIEW").font = Font(bold=True, size=14)
        row += 2
        
        overview_items = [
            ("Client Profile", summary_info['profile'].capitalize()),
            ("Analysis Period", "2009-12 to 2025-12"),
            ("Number of Assets", len(summary_info['allocations'])),
            ("Total Allocation", "100%"),
            ("Data Points", f"{summary_info['n_months']} months")
        ]
        
        for item, value in overview_items:
            ws_summary.cell(row=row, column=1, value=item)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        
        # === SECTION 2: Portfolio Allocations ===
        ws_summary.cell(row=row, column=1, value="PORTFOLIO ALLOCATIONS").font = Font(bold=True, size=14)
        row += 2
        
        ws_summary.cell(row=row, column=1, value="Ticker").font = header_font
        ws_summary.cell(row=row, column=1).fill = header_fill
        ws_summary.cell(row=row, column=2, value="Weight").font = header_font
        ws_summary.cell(row=row, column=2).fill = header_fill
        row += 1
        
        for ticker, weight in sorted(summary_info['allocations'].items(), key=lambda x: -x[1]):
            ws_summary.cell(row=row, column=1, value=ticker)
            ws_summary.cell(row=row, column=2, value=f"{weight*100:.1f}%")
            row += 1
        
        row += 2
        
        # === SECTION 3: Performance Metrics ===
        ws_summary.cell(row=row, column=1, value="PERFORMANCE METRICS").font = Font(bold=True, size=14)
        row += 2
        
        ws_summary.cell(row=row, column=1, value="Metric").font = header_font
        ws_summary.cell(row=row, column=1).fill = header_fill
        ws_summary.cell(row=row, column=2, value="Portfolio").font = header_font
        ws_summary.cell(row=row, column=2).fill = header_fill
        ws_summary.cell(row=row, column=3, value="SPY Benchmark").font = header_font
        ws_summary.cell(row=row, column=3).fill = header_fill
        row += 1
        
        metrics_list = ["Annualized Return", "Annualized Volatility", "Sharpe Ratio", 
                       "Sortino Ratio", "Maximum Drawdown", "Beta", "Jensen's Alpha",
                       "Treynor Ratio", "Information Ratio", "Calmar Ratio"]
        
        def format_val(key, val):
            if key in ["Annualized Return", "Annualized Volatility", "Maximum Drawdown", "Jensen's Alpha"]:
                return f"{float(val)*100:.2f}%"
            elif isinstance(val, (int, float)):
                return f"{float(val):.4f}"
            return str(val)
        
        for m in metrics_list:
            if m in summary_info['port_metrics']:
                ws_summary.cell(row=row, column=1, value=m)
                ws_summary.cell(row=row, column=2, value=format_val(m, summary_info['port_metrics'].get(m, "N/A")))
                ws_summary.cell(row=row, column=3, value=format_val(m, summary_info['bench_metrics'].get(m, "N/A")))
                row += 1
        
        row += 2
        
        # === SECTION 4: Annual Returns ===
        ws_summary.cell(row=row, column=1, value="ANNUAL RETURNS").font = Font(bold=True, size=14)
        row += 2
        
        ws_summary.cell(row=row, column=1, value="Year").font = header_font
        ws_summary.cell(row=row, column=1).fill = header_fill
        ws_summary.cell(row=row, column=2, value="Portfolio").font = header_font
        ws_summary.cell(row=row, column=2).fill = header_fill
        ws_summary.cell(row=row, column=3, value="SPY Benchmark").font = header_font
        ws_summary.cell(row=row, column=3).fill = header_fill
        row += 1
        
        port_annual = summary_info['port_annual']
        bench_annual = summary_info['bench_annual']
        
        for i, date in enumerate(port_annual.index):
            ws_summary.cell(row=row, column=1, value=date.year)
            ws_summary.cell(row=row, column=2, value=f"{port_annual.iloc[i]*100:.2f}%")
            ws_summary.cell(row=row, column=3, value=f"{bench_annual.iloc[i]*100:.2f}%")
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 18
        
        # ========== FORMULA-BASED SHEETS ==========
        # Create "EF Formulas" sheet with Excel formulas referencing Monthly Returns
        ws_formulas = wb.create_sheet("EF Formulas (Linked)")
        
        # Get the monthly returns sheet reference info
        n_months = len(ef_returns)
        n_assets = len(user_tickers)
        last_data_row = n_months + 1  # +1 for header
        
        formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # === Section 1: Asset Statistics with Formulas ===
        ws_formulas['A1'] = "ASSET STATISTICS (EXCEL FORMULAS)"
        ws_formulas['A1'].font = Font(bold=True, size=14)
        ws_formulas.merge_cells('A1:F1')
        
        # Headers
        headers = ['Ticker', 'Mean Monthly Return', 'Mean Annual Return', 'Monthly Std Dev', 'Annual Volatility', 'Sharpe Ratio (Rf=2%)']
        for col, header in enumerate(headers, 1):
            cell = ws_formulas.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Formula rows for each ticker
        for i, ticker in enumerate(user_tickers):
            row = 4 + i
            col_letter = get_column_letter(i + 2)  # Column B, C, D, etc. in Monthly Returns
            data_range = f"'Monthly Returns'!{col_letter}2:{col_letter}{last_data_row}"
            
            ws_formulas.cell(row=row, column=1, value=ticker)
            
            # Mean Monthly Return = AVERAGE(monthly_returns)
            cell_b = ws_formulas.cell(row=row, column=2)
            cell_b.value = f"=AVERAGE({data_range})"
            cell_b.fill = formula_fill
            
            # Mean Annual Return = Monthly * 12
            cell_c = ws_formulas.cell(row=row, column=3)
            cell_c.value = f"=B{row}*12"
            cell_c.fill = formula_fill
            
            # Monthly Std Dev = STDEV(monthly_returns)
            cell_d = ws_formulas.cell(row=row, column=4)
            cell_d.value = f"=STDEV({data_range})"
            cell_d.fill = formula_fill
            
            # Annual Volatility = Monthly * SQRT(12)
            cell_e = ws_formulas.cell(row=row, column=5)
            cell_e.value = f"=D{row}*SQRT(12)"
            cell_e.fill = formula_fill
            
            # Sharpe Ratio = (Annual Return - 2%) / Annual Volatility
            cell_f = ws_formulas.cell(row=row, column=6)
            cell_f.value = f"=(C{row}-0.02)/E{row}"
            cell_f.fill = formula_fill
        
        # === Section 2: Covariance Matrix with Formulas ===
        cov_start_row = 4 + n_assets + 3
        ws_formulas.cell(row=cov_start_row, column=1, value="COVARIANCE MATRIX (ANNUALIZED) - EXCEL FORMULAS")
        ws_formulas.cell(row=cov_start_row, column=1).font = Font(bold=True, size=14)
        
        # Column headers
        for j, ticker in enumerate(user_tickers):
            cell = ws_formulas.cell(row=cov_start_row + 1, column=j + 2, value=ticker)
            cell.font = header_font
            cell.fill = header_fill
        
        # Row headers and covariance formulas
        for i, ticker_i in enumerate(user_tickers):
            row = cov_start_row + 2 + i
            ws_formulas.cell(row=row, column=1, value=ticker_i)
            
            col_letter_i = get_column_letter(i + 2)
            range_i = f"'Monthly Returns'!{col_letter_i}2:{col_letter_i}{last_data_row}"
            
            for j, ticker_j in enumerate(user_tickers):
                col_letter_j = get_column_letter(j + 2)
                range_j = f"'Monthly Returns'!{col_letter_j}2:{col_letter_j}{last_data_row}"
                
                cell = ws_formulas.cell(row=row, column=j + 2)
                # Use COVAR for better Excel compatibility (*12 for annualization)
                cell.value = f"=COVAR({range_i},{range_j})*12"
                cell.fill = formula_fill
        
        # === Section 3: Portfolio Calculator ===
        port_start_row = cov_start_row + n_assets + 5
        ws_formulas.cell(row=port_start_row, column=1, value="PORTFOLIO CALCULATOR")
        ws_formulas.cell(row=port_start_row, column=1).font = Font(bold=True, size=14)
        
        ws_formulas.cell(row=port_start_row + 1, column=1, value="Enter weights below (must sum to 100%):")
        
        # Weight input cells
        headers2 = ['Ticker', 'Weight', 'Expected Return', 'Contribution']
        for col, header in enumerate(headers2, 1):
            cell = ws_formulas.cell(row=port_start_row + 2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for i, ticker in enumerate(user_tickers):
            row = port_start_row + 3 + i
            ws_formulas.cell(row=row, column=1, value=ticker)
            
            # Default weight from current allocation
            curr_weight = allocations.get(ticker, 0)
            ws_formulas.cell(row=row, column=2, value=curr_weight)
            
            # Expected return reference (from Asset Statistics section)
            stats_row = 4 + i
            cell_ret = ws_formulas.cell(row=row, column=3)
            cell_ret.value = f"=C{stats_row}"  # Annual return
            cell_ret.fill = formula_fill
            
            # Contribution = Weight * Return
            cell_contrib = ws_formulas.cell(row=row, column=4)
            cell_contrib.value = f"=B{row}*C{row}"
            cell_contrib.fill = formula_fill
        
        # Summary row
        sum_row = port_start_row + 3 + n_assets
        ws_formulas.cell(row=sum_row, column=1, value="TOTAL:")
        ws_formulas.cell(row=sum_row, column=1).font = Font(bold=True)
        
        # Sum of weights
        weight_range = f"B{port_start_row + 3}:B{sum_row - 1}"
        ws_formulas.cell(row=sum_row, column=2, value=f"=SUM({weight_range})")
        
        # Portfolio Expected Return
        contrib_range = f"D{port_start_row + 3}:D{sum_row - 1}"
        cell_port_ret = ws_formulas.cell(row=sum_row, column=4)
        cell_port_ret.value = f"=SUM({contrib_range})"
        cell_port_ret.fill = formula_fill
        
        # Portfolio Volatility calculation section
        vol_row = sum_row + 2
        ws_formulas.cell(row=vol_row, column=1, value="Portfolio Volatility:")
        ws_formulas.cell(row=vol_row, column=1).font = Font(bold=True)
        
        # Create weight vector reference for MMULT (portfolio variance calculation)
        # For simplicity, we'll use a SUMPRODUCT approach
        ws_formulas.cell(row=vol_row + 1, column=1, value="(Uses SUMPRODUCT with covariance matrix)")
        
        # Sharpe Ratio
        ws_formulas.cell(row=vol_row + 3, column=1, value="Portfolio Sharpe Ratio:")
        ws_formulas.cell(row=vol_row + 3, column=1).font = Font(bold=True)
        ws_formulas.cell(row=vol_row + 3, column=2, value="=(Portfolio Return - 2%) / Volatility")
        
        # === Add Instructions ===
        inst_row = vol_row + 6
        ws_formulas.cell(row=inst_row, column=1, value="INSTRUCTIONS:")
        ws_formulas.cell(row=inst_row, column=1).font = Font(bold=True, size=12)
        
        instructions = [
            "1. All green cells contain Excel formulas that reference the 'Monthly Returns' sheet",
            "2. Modify weights in the Portfolio Calculator to see different portfolio outcomes",
            "3. Weights should sum to 1.0 (100%)",
            "4. Mean Monthly Return = AVERAGE of monthly returns",
            "5. Annual Return = Monthly Return × 12",
            "6. Annual Volatility = Monthly Std Dev × √12",
            "7. Covariance = COVAR × 12 (annualized)",
            "8. Sharpe Ratio = (Annual Return - Risk-Free Rate) / Volatility"
        ]
        
        for i, inst in enumerate(instructions):
            ws_formulas.cell(row=inst_row + 1 + i, column=1, value=inst)
        
        # Adjust column widths
        ws_formulas.column_dimensions['A'].width = 30
        ws_formulas.column_dimensions['B'].width = 20
        ws_formulas.column_dimensions['C'].width = 20
        ws_formulas.column_dimensions['D'].width = 18
        ws_formulas.column_dimensions['E'].width = 18
        ws_formulas.column_dimensions['F'].width = 20
        
        # ========== CREATE EXCEL-NATIVE SCATTER CHART ==========
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.label import DataLabelList
        
        # Get the Efficient Frontier sheet
        ws_ef = wb["Efficient Frontier"]
        ef_data_rows = ws_ef.max_row
        
        # ========== ADD CAL DATA TO EXCEL ==========
        # Re-calculate CAL parameters to be safe
        rf_rate = 0.02
        tangency_vol = max_sharpe_portfolio['Portfolio Std Dev']
        tangency_ret = max_sharpe_portfolio['Portfolio Expected Returns']
        cal_slope = (tangency_ret - rf_rate) / tangency_vol
        
        max_vol_pf = ef_df['Portfolio Std Dev'].max()
        cal_end_vol = max_vol_pf * 1.2
        cal_end_ret = rf_rate + cal_slope * cal_end_vol
        
        # Write CAL data to columns (leaving a gap of 2 columns)
        n_tickers = len(user_tickers)  # Define this here!
        
        # Columns: [Weights] | StdDev | Ret | Sharpe
        # StdDev is col n_tickers + 1
        # Ret is n_tickers + 2
        # Sharpe is n_tickers + 3
        
        cal_col_vol = n_tickers + 6
        cal_col_ret = n_tickers + 7
        
        ws_ef.cell(row=1, column=cal_col_vol, value="CAL Vol")
        ws_ef.cell(row=1, column=cal_col_ret, value="CAL Return")
        
        # Point 1 (Risk free)
        ws_ef.cell(row=2, column=cal_col_vol, value=0)
        ws_ef.cell(row=2, column=cal_col_ret, value=rf_rate)
        
        # Point 2 (Extended)
        ws_ef.cell(row=3, column=cal_col_vol, value=cal_end_vol)
        ws_ef.cell(row=3, column=cal_col_ret, value=cal_end_ret)
        
        # Create scatter chart
        chart = ScatterChart()
        chart.title = "Efficient Frontier (Excel-Linked)"
        chart.x_axis.title = "Portfolio Std Dev"
        chart.y_axis.title = "Portfolio Expected Returns"
        chart.style = 10
        chart.width = 18
        chart.height = 12
        
        # Column layout
        # n_tickers = len(user_tickers) # Already defined above
        std_dev_col = n_tickers + 1    # Portfolio Std Dev column
        exp_ret_col = n_tickers + 2    # Portfolio Expected Returns column
        
        # X values (Portfolio Std Dev)
        x_values = Reference(ws_ef, min_col=std_dev_col, min_row=2, max_row=ef_data_rows)
        # Y values (Portfolio Expected Returns)
        y_values = Reference(ws_ef, min_col=exp_ret_col, min_row=2, max_row=ef_data_rows)
        
        # Series 1: Random Portfolios (Blue Dots)
        series = Series(y_values, x_values, title="Portfolios")
        series.marker = Marker(symbol='circle', size=5)
        series.graphicalProperties.line.noFill = True  # No connecting lines
        # Set color to Blue (#4B87B9)
        series.marker.graphicalProperties.solidFill = "4B87B9"
        series.marker.graphicalProperties.line.solidFill = "4B87B9" 
        chart.series.append(series)
        
        # Series 2: CAL (Orange Line)
        cal_x = Reference(ws_ef, min_col=cal_col_vol, min_row=2, max_row=3)
        cal_y = Reference(ws_ef, min_col=cal_col_ret, min_row=2, max_row=3)
        
        series_cal = Series(cal_y, cal_x, title="CAL")
        series_cal.marker.symbol = "none" # No marker
        series_cal.graphicalProperties.line.solidFill = "FFA500" # Orange
        series_cal.graphicalProperties.line.dashStyle = "sysDash" # Dashed line
        chart.series.append(series_cal)
        
        # Create chart sheet
        ws_chart = wb.create_sheet("EF Chart (Excel)")
        ws_chart.add_chart(chart, "A1")
        
        # Also keep the PNG chart for reference
        ws_png = wb.create_sheet("EF Chart (Image)")
        img = Image(chart_path)
        img.width = 800
        img.height = 533
        ws_png.add_image(img, 'A1')
        
        wb.save(output_path)
        
        return jsonify({
            "success": True, 
            "filename": filename,
            "chart": chart_filename
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    """Serve the Excel file for download."""
    return send_from_directory('output', filename, as_attachment=True)



if __name__ == '__main__':
    # Create output directory
    os.makedirs('output', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    print("\n" + "="*60)
    print("  FNCE102 Portfolio Analyzer - Web Interface")
    print("="*60)
    print("\n  Open your browser and go to: http://127.0.0.1:5000")
    print("\n  Press Ctrl+C to stop the server")
    print("="*60 + "\n")
    
    app.run(debug=True, port=5000)
